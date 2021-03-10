
import time
import urllib.request  as urllib2 
from time      import sleep
from selenium  import webdriver
from multiprocessing import Pool
from itertools import product
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent
from openpyxl import load_workbook
import sys
from selenium.webdriver.common.action_chains import ActionChains


processName = {}
pins_per_layer = 2
pin_count = 0

      
def logIn(mail, url, driver): #, proxy_ip) :
    print("logging in...")
    googleLogin(mail, driver)


def fail_with_error(message):
    def decorator(fx):
        def inner(*args, **kwargs):
            try:
                return fx(*args, **kwargs)
            except Exception as e:
                print(message)
                raise e
        return inner
    return decorator

@fail_with_error("Cannot set email address")
def google_set_login(driver, mail_address):
    try:
        email_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="identifierId"]')))
        email_field.send_keys(mail_address)
        print("Email address inserted")
    except TimeoutException:
        print("email field is not ready")
        pass

@fail_with_error("Cannot click login button")
def google_click_login_button(driver):
    try:
        login_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="identifierNext"]')))
        login_button.click()
        print("Login button clicked")
    except TimeoutException:
        print("login button is not ready")
        pass

@fail_with_error("Cannot set email password")
def google_set_password(driver, password):
    try:
        password_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input')))
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input')))
        password_field.send_keys(password)
        print("Password inserted")
    except TimeoutException:
        print("password field is not ready")
        pass

@fail_with_error("Cannot click next button")
def google_click_next_button(driver):
    try:
        next_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="passwordNext"]')))
        next_button.click()
        print("Next button clicked")
    except TimeoutException:
        print("next button is not ready")
        pass


def googleLogin(mail, driver) :
    print("gmail logging in...")
    driver.get('https://accounts.google.com/signin/v2/identifier?passive=1209600&continue=https%3A%2F%2Fwww.google.com%2Fmaps%2Fd%2F&followup=https%3A%2F%2Fwww.google.com%2Fmaps%2Fd%2F&flowName=GlifWebSignIn&flowEntry=ServiceLogin')
    mail_address = mail[1]
    mail_pass = mail[2]
    mail_active = mail[3]
    print("opend browser")
    google_set_login(driver, mail_address)
    google_click_login_button(driver)
    google_set_password(driver, mail_pass)
    google_click_next_button(driver)


@fail_with_error("Cannot click 'Create a new map' button")
def create_map(driver):
    try:
        crate_map_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//span[text()='+ Create a new map']/parent::span/parent::div")))
        crate_map_btn.click()
        print("'Crate a new map' button clicked")
    except TimeoutException:
        print("'Crate a new map' button is not ready")
        pass


# @fail_with_error("Cannot import data")
# def import_data(driver):
#     try:
#         import_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@data-tooltip='Import data from a CSV file, spreadsheet or KML.']")))
#         import_btn.click()
#         print("Import button clicked")
#     except TimeoutException:
#         print("Import button is not ready")
#         return

#   time.sleep(10)

#   try:
    
#     # driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
#     # driver.switch_to_frame(driver.find_element_by_css_selector("body>iframe"))
#     # iframe = driver.find_elements_by_tag_name('iframe')[5]
#     driver.switch_to.frame(5)
#     # driver.switchTo().frame(driver.findElement(By.name("iFrameTitle")));

#     google_drive_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[text()='Google Drive']")))
#     google_drive_btn.click()
#     print("'Google drive' button clicked")
#   except TimeoutException:
#     print("'Google drive' button is not ready")
#     return

#   time.sleep(10)

#   try:
#     sel_doc = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[contains(@data-id,'1-bwl7IqGW2FhNvphO8gP1MrvZXhY0PWD6aVaFNDUkvk') and not(@id)]")))
#     sel_doc.click()
#     print("doc is selected.")
#   except TimeoutException:
#     print("doc is not ready")
#     return
  
#   time.sleep(10)
#   try:
#     select_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "/html/body/div[2]/div/div[4]/div[2]/div/div[2]/div/div[2]//div[@role='button' and text()='Select']")))
#     select_btn.click()
#     print("select_btn is selected.")
#   except TimeoutException:
#     print("select_btn is not ready")
#     return

#   driver.switch_to.default_content()
#   time.sleep(20)
#   try:
#     coor_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "/html//span[contains(@id, 'upload-checkbox-label') and text()='Coordenadas']"))) # //span[contains(@id, 'upload-checkbox-label') and text()='Coordenadas'] //preceding-sibling::span/span]
#     # /html/body/div[9]/div[2]/div/div[3]/div[4]/span[2]
#     coor_btn.click()
#     print("coor_btn is selected.")
#   except TimeoutException:
#     print("coor_btn is not ready")
#     return

#   time.sleep(20)
#   try:
#     long_lat_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Longitude,Latitude']")))
#     long_lat_btn.click()
#     print("long_lat_btn is selected.")
#   except TimeoutException:
#     print("long_lat_btn is not ready")
#     return

#   time.sleep(10)
#   try:
#     continue_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//button[text()='Continue']")))
#     continue_btn.click()
#     print("continue_btn is selected.")
#   except TimeoutException:
#     print("continue_btn is not ready")
#     return

#   time.sleep(10)
#   try:
#     map_title_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//span[text()='Business Name' and not(@id)]")))
#     map_title_btn.click()
#     print("map_title_btn is selected.")
#   except TimeoutException:
#     print("map_title_btn is not ready")
#     return

#   time.sleep(10)
#   try:
#     finish_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//button[text()='Finish']")))
#     finish_btn.click()
#     print("finish_btn is selected.")
#   except TimeoutException:
#     print("finish_btn is not ready")
#     return

#   # time.sleep(5)
#   # try:
#   #   continue_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//button[text()='Continue']")))
#   #   continue_btn.click()
#   #   print("continue_btn is selected.")
#   # except TimeoutException:
#   #   print("continue_btn is not ready")
#   #   return
  

#   # try:
#   #   import_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@data-tooltip='Import data from a CSV file, spreadsheet or KML.']")))
#   #   import_btn.click()
#   #   print("'Import' button clicked")
#   # except TimeoutException:
#   #   print("'Import' button is not ready")
#   #   return


@fail_with_error("Cannot create pins")
def create_pins(driver, file_name, coordinate_col_name, text_col_name):
    create_pins_from_excel(driver, file_name, coordinate_col_name, text_col_name)
    # create_pin(driver, 5.55555, 6.66666, "111")
    # create_pin(driver, 35.55555, 36.66666, "222")


def create_pins_from_excel(driver, file_name, coordinate_col_name, text_col_name):
    wb = load_workbook(filename = file_name)
    sheet = wb.active
    col_num = -1
    coordinate_col_num = -1
    text_col_num = -1
    print(coordinate_col_name)
    print(text_col_name)

    for col in sheet[1]:
        col_num += 1
        print(col.value)
        if col.value == coordinate_col_name: coordinate_col_num = col_num
        if col.value == text_col_name: text_col_num = col_num

    if coordinate_col_num == -1 or text_col_num == -1: 
        print("Wrong Column Names")
        exit()

    for row in sheet[2:sheet.max_row]:
        coordinate = row[coordinate_col_num].value
        if coordinate == None : continue
        if coordinate.strip() == "": continue
        text = row[text_col_num].value.strip()
        print(coordinate + " :: " + text)
        lat_lng = coordinate.split(",")
        create_pin(driver, float(lat_lng[0]), float(lat_lng[1]), text)


@fail_with_error("Cannot click marker button")
def marker_click(driver):
    try:
        marker_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@id='markerButton']")))
        marker_btn.click()
        print("marker_btn button clicked")
    except TimeoutException:
        print("marker_btn button is not ready")
        pass
    time.sleep(1)


def name_map(driver, text):
    try:
        untitled_map_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@data-tooltip='Untitled map']")))
        untitled_map_btn.click()
        print("untitled_map_btn button clicked")
    except TimeoutException:
        print("untitled_map_btn button is not ready")
        pass


    try:
        title_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@type='text' and @placeholder='Untitled map']")))
        title_field.clear()
        title_field.send_keys(text)
        print("title field is filled.")
    except TimeoutException:
        print("title field is not ready")
        pass

    try:
        desc_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//textarea[@placeholder='Add a description to help people understand your map']")))
        desc_field.clear()
        desc_field.send_keys(text)
        print("description field is filled.")
    except TimeoutException:
        print("description field is not ready")
        pass

    try:
        save_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@id='update-map']//button[@name='save']")))
        save_btn.click()
        print("save button clicked")
    except TimeoutException:
        print("save button is not ready")
        pass
    time.sleep(2)


def name_layer(driver, text):
    try:
        untitled_layer_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@data-tooltip='Untitled layer']")))
        untitled_layer_btn.click()
        print("untitled_layer_btn button clicked")
    except TimeoutException:
        print("untitled_layer_btn button is not ready")
        pass

    try:
        title_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@type='text' and @value='Untitled layer']")))
        title_field.clear()
        title_field.send_keys(text)
        print("title field is filled.")
    except TimeoutException:
        print("title field is not ready")
        pass

    try:
        save_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@id='update-layer-name']//button[@name='save']")))
        save_btn.click()
        print("save button clicked")
    except TimeoutException:
        print("save button is not ready")
        pass
    time.sleep(2)


def add_layer(driver):
    try:
        add_layer_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//li[@id='map-action-add-layer']")))
        add_layer_btn.click()        
        print("add_layern button clicked")
    except TimeoutException:
        print("add_layer button is not ready")
        pass
    time.sleep(2)


@fail_with_error("Cannot create pin")
def create_pin(driver, lat, lng, text): 
    global pin_count, pins_per_layer
    pin_count += 1
    if pin_count == 1: name_map(driver, text)
    if pin_count % pins_per_layer == 1: 
        if (pin_count - 1) // pins_per_layer > 0: add_layer(driver)
        name_layer(driver, text)
    print(driver.current_url) 
    url_segs = driver.current_url.split("&")
    if lat > 0 : 
        lat += 0.000005 
    else:
        lat -= 0.000005
    if lng > 0 : 
        lng += 0.000005 
    else:
        lng -= 0.000005

    url_segs[-2] = "ll=" + str(lat) + "%2C" +str(lng)
    url_segs[-1] = "z=22"
    try:
        driver.get("&".join(url_segs))
        time.sleep(3)
        sel_layer = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "(//div[contains(@id, 'layer-header')])[last()]/parent::div")))
        sel_layer.click()
        marker_click(driver)
        action = ActionChains(driver)
        elem_origin = driver.find_element_by_xpath('//html')
        print(elem_origin.rect)
        bounds = elem_origin.size
        print (bounds)
        
        action.move_to_element_with_offset(elem_origin, bounds['width'] / 2, bounds['height'] / 2)
        time.sleep(2)
        print(str(bounds['width'] / 2) + " , " + str(bounds['height'] / 2))
        action.click()
        print("clicked")
        action.perform()
        print("performed")
        time.sleep(2)

        title_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@id='map-infowindow-attr-name-value']")))
        title_field.clear()
        title_field.send_keys(text)
        time.sleep(2)

        desc_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@id='map-infowindow-attr-description-value']")))
        desc_field.send_keys(text)
        time.sleep(2)

        save_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@role='button' and text()='Save']")))
        save_btn.click()
        print("save_btn clicked")

        print("Crated a pin")
        time.sleep(2)

    # marker_click(driver)
    # action = ActionChains(driver)
    # # action.move_to_element_with_offset(elem_body, 0, 0).perform()
    # # time.sleep(2)
    # action.move_to_element_with_offset(elem_origin, 592, 384)
    # time.sleep(5)
    # print(str(bounds['width'] / 2) + " , " + str(bounds['height'] / 2))
    # action.click()
    # action.perform()
    # time.sleep(5)
    # save_btn = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//div[@role='button' and text()='Save']")))
    # save_btn.click()
    # print("Crated a pin")

    # url_segs = driver.current_url.split("&")
    # url_segs[-2] = "ll=9.9999%2C9.9999"
    # url_segs[-1] = "z=4"
    # driver.get("&".join(url_segs))
    except TimeoutException:
        print("Cannot create pin")
        pass


# work_log(["1", 'janwheeler197335@gmail.com', 'aleiivrc', 'nwxdrjem'])

# def work_log(mail):
if len(sys.argv) < 4:
    print("python 3.py [file name] [coordinate column name] [text column name]")
    exit()

file_name = sys.argv[1]
coordinate_col_name = sys.argv[2]
text_col_name = sys.argv[3]

url = 'https://accounts.google.com/'
mail = ["1", 'janwheeler197335@gmail.com', 'aleiivrc', 'nwxdrjem']

ua = UserAgent()
userAgent = ua.random
userAgent = userAgent.split(" ")
# userAgent[0] = "Mozilla/5.0"
userAgent = " ".join(userAgent)
print("userAgent = " + userAgent)
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('user-agent={0}'.format(userAgent))
path = '.\\webdriver\\chromedriver.exe'
chrome_options.add_argument('--log-level=0')
driver = webdriver.Chrome (executable_path = path, options = chrome_options )
# driver.maximize_window()
driver.set_window_size(1200,900)

try:
    processName[mail[0]] = logIn
    processName[mail[0]](mail, url, driver) #, proxies)
    time.sleep(10)
    create_map(driver)
    time.sleep(10)
    create_pins(driver, file_name, coordinate_col_name, text_col_name)
except Exception as e:
    # driver.save_screenshot(datetime.now().strftime("screenshot_%Y%m%d_%H%M%S_%f.png"))
    print(e)
    # driver.quit()
finally:
    pass

  # //div[text()='Select a file from your device']
  # //div[text()='Google Drive']
  # https://drive.google.com/file/d/1wabQB2NV19BZR4jAeZN15Fl8-eMgTUZY/view?usp=sharing
  # https://docs.google.com/spreadsheets/d/1-bwl7IqGW2FhNvphO8gP1MrvZXhY0PWD6aVaFNDUkvk/edit?usp=sharing
  #data-id :2s.ritz.0.1-bwl7IqGW2FhNvphO8gP1MrvZXhY0PWD6aVaFNDUkvk

  # <div id=":2s.ritz.0.1-bwl7IqGW2FhNvphO8gP1MrvZXhY0PWD6aVaFNDUkvk" data-id="ritz.0.1-bwl7IqGW2FhNvphO8gP1MrvZXhY0PWD6aVaFNDUkvk" data-target="doc" draggable="true" role="option" tabindex="0" aria-selected="false" aria-disabled="false"><div class="Nd-eg-qg" data-target="itemUploadDrop"><div class="Nd-eg-qg-Ag-qb Nd-eg-Od-Oc"><div class="Nd-dg-Fe-Dg"><svg height="100%" width="100%" viewBox="0 0 10 10" preserveAspectRatio="none" focusable="false"><rect width="10" height="10"></rect></svg></div><div class="Nd-eg-qg-rg"><div class="Od-eg-Oc-Kc" data-id="ritz.0.1-bwl7IqGW2FhNvphO8gP1MrvZXhY0PWD6aVaFNDUkvk" aria-label="Juegos de Poniente Madrid Shared Google Sheets"><div class="Od-eg-Oc-f-Kc" data-is-doc-name="true"><div class="Od-eg-f-eb"><div class="Od-eg-f Od-eg-lh-sc Od-eg-zh"><img src="//ssl.gstatic.com/docs/doclist/images/mediatype/icon_1_spreadsheet_x16.png" alt=""></div><div class="Od-eg-f Od-eg-lh-sc"><img src="//ssl.gstatic.com/docs/doclist/images/mediatype/icon_1_spreadsheet_x16.png" alt=""></div></div></div><div class="Od-eg-Oc-cb" data-tooltip="Google Sheets: Juegos de Poniente Madrid"><span class="Od-Ff-Oc-tc" data-is-doc-name="true">Juegos de Poniente Madrid</span></div><div class="Od-eg-Gh-ie-Xd-Kc"></div><div class="Od-eg-Ie-Kc"><div class="Od-eg-b-f Od-eg-Ie" data-tooltip="Shared" aria-label="Shared" data-tooltip-align="b,c" data-tooltip-delay="1000" data-tooltip-unhoverable="true"><div class="Od-eg-b-f-Ph"><svg class="undefined Nd-te-I-lf-Ae" width="16px" height="16px" viewBox="0 0 16 16" focusable="false" fill="#000000"><path d="M5,7 C6.11,7 7,6.1 7,5 C7,3.9 6.11,3 5,3 C3.9,3 3,3.9 3,5 C3,6.1 3.9,7 5,7 L5,7 Z M11,7 C12.11,7 13,6.1 13,5 C13,3.9 12.11,3 11,3 C9.89,3 9,3.9 9,5 C9,6.1 9.9,7 11,7 L11,7 Z M5,8.2 C3.33,8.2 0,9.03 0,10.7 L0,12 L10,12 L10,10.7 C10,9.03 6.67,8.2 5,8.2 L5,8.2 Z M11,8.2 C10.75,8.2 10.46,8.22 10.16,8.26 C10.95,8.86 11.5,9.66 11.5,10.7 L11.5,12 L16,12 L16,10.7 C16,9.03 12.67,8.2 11,8.2 L11,8.2 Z"></path></svg></div><div class="Od-eg-b-f-w"><svg class="undefined Nd-te-I-lf-Ae" width="16px" height="16px" viewBox="0 0 16 16" focusable="false" fill="#FFFFFF"><path d="M5,7 C6.11,7 7,6.1 7,5 C7,3.9 6.11,3 5,3 C3.9,3 3,3.9 3,5 C3,6.1 3.9,7 5,7 L5,7 Z M11,7 C12.11,7 13,6.1 13,5 C13,3.9 12.11,3 11,3 C9.89,3 9,3.9 9,5 C9,6.1 9.9,7 11,7 L11,7 Z M5,8.2 C3.33,8.2 0,9.03 0,10.7 L0,12 L10,12 L10,10.7 C10,9.03 6.67,8.2 5,8.2 L5,8.2 Z M11,8.2 C10.75,8.2 10.46,8.22 10.16,8.26 C10.95,8.86 11.5,9.66 11.5,10.7 L11.5,12 L16,12 L16,10.7 C16,9.03 12.67,8.2 11,8.2 L11,8.2 Z"></path></svg></div></div></div><div class="Od-eg-Rf-Nd-Kc"></div><div class="Od-eg-Rf-Nd-Hc-Sh-Kc"><div></div></div><div class="Od-eg-zg-He-Kc"><div></div></div><div class="Od-eg-Qh-He-Kc"><div></div></div><div class="Od-eg-Rh-Kc"><div></div></div><div class="Od-eg-Th-Uh-Kc"><div></div></div><div class="Od-eg-Nh-Ne-Kc"><div class="Od-eg-b-f Od-eg-Nh-Ne"></div></div><div class="Od-eg-zg-Kc"></div><div class="Od-eg-Bh-Kc"></div></div></div></div></div><div class="Nd-eg-qg"><div class="Nd-eg-qg-Ag-qb Nd-eg-Od-Re"><div class="Nd-dg-Fe-Dg"><svg height="100%" width="100%" viewBox="0 0 10 10" preserveAspectRatio="none" focusable="false"><rect width="10" height="10"></rect></svg></div><div class="Nd-eg-qg-rg"><div class="Nd-te-Le-Kc"><span class="Nd-te-Le-ye">me</span></div></div></div></div><div class="Nd-eg-qg"><div class="Nd-eg-qg-Ag-qb Nd-eg-Od-Wg"><div class="Nd-dg-Fe-Dg"><svg height="100%" width="100%" viewBox="0 0 10 10" preserveAspectRatio="none" focusable="false"><rect width="10" height="10"></rect></svg></div><div class="Nd-eg-qg-rg"><div class="Nd-te-Le-Kc" data-tooltip-align="b,c" data-tooltip-delay="500" data-tooltip-unhoverable="true" data-tooltip="Last modified 8:04 AM" aria-label="Last modified 8:04 AM"><span class="Nd-te-Le-ye">8:04 AM</span><span class="Nd-te-Le-Vd"></span></div></div></div></div></div>
  # <div data-id="ritz.0.1-bwl7IqGW2FhNvphO8gP1MrvZXhY0PWD6aVaFNDUkvk" aria-label="Juegos de Poniente Madrid Shared Google Sheets"><img src="//ssl.gstatic.com/docs/doclist/images/mediatype/icon_1_spreadsheet_x16.png" alt=""></div><div class="Od-eg-f Od-eg-lh-sc"><img src="//ssl.gstatic.com/docs/doclist/images/mediatype/icon_1_spreadsheet_x16.png" alt=""></div></div></div><div class="Od-eg-Oc-cb" data-tooltip="Google Sheets: Juegos de Poniente Madrid"><span class="Od-Ff-Oc-tc" data-is-doc-name="true">Juegos de Poniente Madrid</span></div><div class="Od-eg-Gh-ie-Xd-Kc"></div><div class="Od-eg-Ie-Kc"><div class="Od-eg-b-f Od-eg-Ie" data-tooltip="Shared" aria-label="Shared" data-tooltip-align="b,c" data-tooltip-delay="1000" data-tooltip-unhoverable="true"><div class="Od-eg-b-f-Ph"><svg class="undefined Nd-te-I-lf-Ae" width="16px" height="16px" viewBox="0 0 16 16" focusable="false" fill="#000000"><path d="M5,7 C6.11,7 7,6.1 7,5 C7,3.9 6.11,3 5,3 C3.9,3 3,3.9 3,5 C3,6.1 3.9,7 5,7 L5,7 Z M11,7 C12.11,7 13,6.1 13,5 C13,3.9 12.11,3 11,3 C9.89,3 9,3.9 9,5 C9,6.1 9.9,7 11,7 L11,7 Z M5,8.2 C3.33,8.2 0,9.03 0,10.7 L0,12 L10,12 L10,10.7 C10,9.03 6.67,8.2 5,8.2 L5,8.2 Z M11,8.2 C10.75,8.2 10.46,8.22 10.16,8.26 C10.95,8.86 11.5,9.66 11.5,10.7 L11.5,12 L16,12 L16,10.7 C16,9.03 12.67,8.2 11,8.2 L11,8.2 Z"></path></svg></div><div class="Od-eg-b-f-w"><svg class="undefined Nd-te-I-lf-Ae" width="16px" height="16px" viewBox="0 0 16 16" focusable="false" fill="#FFFFFF"><path d="M5,7 C6.11,7 7,6.1 7,5 C7,3.9 6.11,3 5,3 C3.9,3 3,3.9 3,5 C3,6.1 3.9,7 5,7 L5,7 Z M11,7 C12.11,7 13,6.1 13,5 C13,3.9 12.11,3 11,3 C9.89,3 9,3.9 9,5 C9,6.1 9.9,7 11,7 L11,7 Z M5,8.2 C3.33,8.2 0,9.03 0,10.7 L0,12 L10,12 L10,10.7 C10,9.03 6.67,8.2 5,8.2 L5,8.2 Z M11,8.2 C10.75,8.2 10.46,8.22 10.16,8.26 C10.95,8.86 11.5,9.66 11.5,10.7 L11.5,12 L16,12 L16,10.7 C16,9.03 12.67,8.2 11,8.2 L11,8.2 Z"></path></svg></div></div></div><div class="Od-eg-Rf-Nd-Kc"></div><div class="Od-eg-Rf-Nd-Hc-Sh-Kc"><div></div></div><div class="Od-eg-zg-He-Kc"><div></div></div><div class="Od-eg-Qh-He-Kc"><div></div></div><div class="Od-eg-Rh-Kc"><div></div></div><div class="Od-eg-Th-Uh-Kc"><div></div></div><div class="Od-eg-Nh-Ne-Kc"><div class="Od-eg-b-f Od-eg-Nh-Ne"></div></div><div class="Od-eg-zg-Kc"></div><div class="Od-eg-Bh-Kc"></div></div>

  # /html/body/div[2]/div/div[4]/div[2]/div/div[2]/div/div[2]//div[@role='button' and text()='Select']
  # //span[contains(@id, 'upload-checkbox-label') and text()='Coordenadas']//preceding-sibling::span/span  
  # //span[text()='Longitude,Latitude']
  # //button[text()='Continue']
  # //span[text()='Business Name' and not(@id)]
  # //button[text()='Finish']


  # //div[@id='map-infowindow-attr-name-value']
  # //div[@id='map-infowindow-attr-description-value']
  # //div[@role='button' and text()='Save']


  # //div[@data-tooltip='Untitled layer']
  # //input[@type='text' and @value='Untitled layer']
  # //button[@name='save']


#   //li[@id='map-action-add-layer']
