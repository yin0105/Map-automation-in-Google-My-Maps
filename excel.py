from openpyxl import load_workbook
import sys

wb = load_workbook(filename = 'Juegos de Poniente Madrid.xlsx')
# sheet_ranges = wb['Hoja 1']
# print(sheet_ranges['D18'].value)

sheet = wb.active

# for row in sheet.rows:
#     # print(row)
#     for col in row:
#         # print(col)
#         pass
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet[1])
# for col in sheet[1]:
#     print(col)

if len(sys.argv) < 4:
    print("python 3.py [file name] [coordinate column name] [text column name]")
    exit()

file_name = sys.argv[1]
coordinate_col_name = sys.argv[2]
text_col_name = sys.argv[3]
col_num = -1
coordinate_col_num = -1
text_col_num = -1
print(coordinate_col_name)
print(text_col_name)

for col in sheet[1]:
    col_num += 1
    print(col.value)
    if col.value == coordinate_col_name: coordinate_col_num = col_num
    if col.value == text_col_name: text_col_num = col_num

if coordinate_col_num == -1 or text_col_num == -1: 
    print("Wrong Column Names")
    exit()

for row in sheet[2:sheet.max_row]:
    coordinate = row[coordinate_col_num].value
    if coordinate == None : continue
    if coordinate.strip() == "": continue
    text = row[text_col_num].value.strip()
    print(coordinate + " :: " + text)

