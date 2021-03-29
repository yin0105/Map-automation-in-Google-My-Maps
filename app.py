
import time, sys, os, pickle
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent
from openpyxl import load_workbook
from selenium.webdriver.common.action_chains import ActionChains
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from os.path import join, dirname
from dotenv import load_dotenv


dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)
pins_per_layer = int(os.environ.get('PINS_PER_LAYER'))
layer_count = int(os.environ.get('LAYER_COUNT'))

processName = {}
pin_count = 0

      
def logIn(mail, driver):
    print("logging in...")
    googleLogin(mail, driver)


def fail_with_error(message):
    def decorator(fx):
        def inner(*args, **kwargs):
            try:
                return fx(*args, **kwargs)
            except Exception as e:
                print(message)
                raise e
        return inner
    return decorator


@fail_with_error("Cannot set email address")
def google_set_login(driver, mail_address):
    while True:
        try:
            email_field = driver.find_element_by_xpath("//*[@id='identifierId']")
            email_field.send_keys(mail_address)
            print("Email address inserted")
            break
        except :
            time.sleep(1)

    # try:
    #     email_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="identifierId"]')))
    #     email_field.send_keys(mail_address)
    #     print("Email address inserted")
    # except TimeoutException:
    #     try:
    #         email_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@id='Email']")))
    #         email_field.send_keys(mail_address)
    #         print("Email address inserted")
    #     except TimeoutException:
    #         print("email field is not ready")
    

@fail_with_error("Cannot click login button")
def google_click_login_button(driver):
    while True:
        try:
            login_button = driver.find_element_by_xpath("//*[@id='identifierNext']")
            login_button.click()
            print("Login button clicked")
            break
        except :
            time.sleep(1)

    # try:
    #     login_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="identifierNext"]')))
    #     login_button.click()
    #     print("Login button clicked")
    # except TimeoutException:
    #     try:
    #         login_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//input[@id='next']")))
    #         login_button.click()
    #         print("Login button clicked")
    #     except TimeoutException:
    #         print("login button is not ready")


@fail_with_error("Cannot set email password")
def google_set_password(driver, password):
    while True:
        try:
            password_field = driver.find_element_by_xpath("//*[@id='password']/div[1]/div/div[1]/input")
            # WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input')))
            password_field.send_keys(password)
            print("Password inserted")
            break
        except :
            time.sleep(1)

    # try:
    #     password_field = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input')))
    #     WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input')))
    #     password_field.send_keys(password)
    #     print("Password inserted")
    # except TimeoutException:
    #     print("password field is not ready")
    #     pass


@fail_with_error("Cannot click next button")
def google_click_next_button(driver):
    while True:
        try:
            next_button = driver.find_element_by_xpath("//*[@id='passwordNext']")
            next_button.click()
            print("Next button clicked")
            break
        except :
            time.sleep(1)

    # try:
    #     next_button = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="passwordNext"]')))
    #     next_button.click()
    #     print("Next button clicked")
    # except TimeoutException:
    #     print("next button is not ready")
    #     pass


def googleLogin(mail, driver) :
    print("gmail logging in...")
    driver.get('https://accounts.google.com/signin/v2/identifier?passive=1209600&continue=https%3A%2F%2Fwww.google.com%2Fmaps%2Fd%2F&followup=https%3A%2F%2Fwww.google.com%2Fmaps%2Fd%2F&flowName=GlifWebSignIn&flowEntry=ServiceLogin')
    mail_address = mail[0]
    mail_pass = mail[1]
    # mail_active = mail[2]
    print("opend browser")
    google_set_login(driver, mail_address)
    google_click_login_button(driver)
    google_set_password(driver, mail_pass)
    google_click_next_button(driver)


@fail_with_error("Cannot click 'Create a new map' button")
def create_map(driver):
    try:
        crate_map_btn = driver.find_element_by_xpath("//span[text()='+ Create a new map']/parent::span/parent::div")
        crate_map_btn.click()
        print("'Crate a new map' button clicked")
    except TimeoutException:
        print("'Crate a new map' button is not ready")
        pass


@fail_with_error("Cannot create pins")
def create_pins(driver, file_name, coordinate_col_name, text_col_name):
    file_name_segs = file_name.split(".")
    if len(file_name_segs) > 1 and file_name_segs[-1][:3].lower() == "xls":
        create_pins_from_excel(driver, file_name, coordinate_col_name, text_col_name)
    else:
        create_pins_from_google_sheet(driver, file_name, coordinate_col_name, text_col_name)


def create_pins_from_excel(driver, file_name, coordinate_col_name, text_col_name):
    wb = load_workbook(filename = file_name)
    sheet = wb.active
    col_num = -1
    coordinate_col_num = -1
    text_col_num = -1
    print(coordinate_col_name)
    print(text_col_name)

    for col in sheet[1]:
        col_num += 1
        print(col.value)
        if col.value == coordinate_col_name: coordinate_col_num = col_num
        if col.value == text_col_name: text_col_num = col_num

    if coordinate_col_num == -1 or text_col_num == -1: 
        print("Wrong Column Names")
        exit()

    for layer_num in range(0, layer_count):
        count = 0
        for row in sheet[2:sheet.max_row]:
            coordinate = row[coordinate_col_num].value
            if coordinate == None : continue
            if coordinate.strip() == "": continue
            text = row[text_col_num].value.strip()
            print(coordinate + " :: " + text)
            lat_lng = coordinate.split(",")
            count += 1
            if count > pins_per_layer: break
            create_pin(driver, float(lat_lng[0]), float(lat_lng[1]), text)


def create_pins_from_google_sheet(driver, file_name, coordinate_col_name, text_col_name):
    col_num = -1
    coordinate_col_num = -1
    text_col_num = -1
    print(coordinate_col_name)
    print(text_col_name)

    creds = None

    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
                creds = flow.run_local_server(port=21000)
            with open('token.pickle', 'wb') as token:
                pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)

    sheet_id =  file_name
    sheet_metadata = service.spreadsheets().get(spreadsheetId=sheet_id).execute()
    sheets = sheet_metadata.get('sheets', '')
    sheet_name =  sheets[0]['properties']['title']
    print(sheet_name)
    sheet_range =  "A1:ZZ100000"
    # Call the Sheets API
    # sheet = service.spreadsheets()
    result = service.spreadsheets().values().get(spreadsheetId=sheet_id, range=sheet_name + "!" + sheet_range).execute()

    values = result.get('values', [])
    resp = ""
    if values:
        # Get Header 
        for col in values[0]:
            col_num += 1
            if col == coordinate_col_name: coordinate_col_num = col_num
            if col == text_col_name: text_col_num = col_num
        
        if coordinate_col_num == -1 or text_col_num == -1: 
            print("Wrong Column Names")
            exit()

        for layer_num in range(0, layer_count):
            count = 0
            for row in values[1:]:
                coordinate = row[coordinate_col_num]
                if coordinate == None : continue
                if coordinate.strip() == "": continue
                text = row[text_col_num].strip()
                print(coordinate + " :: " + text)
                lat_lng = coordinate.split(",")
                count += 1
                if count > pins_per_layer: break
                create_pin(driver, float(lat_lng[0]), float(lat_lng[1]), text)
            
    else:
        print("Wrong Google Sheet Info")
        exit()
        


@fail_with_error("Cannot click marker button")
def marker_click(driver):
    while True:
        try:
            marker_btn = driver.find_element_by_xpath("//div[@id='markerButton']")
            marker_btn.click()
            print("marker_btn button clicked")
            break
        except :
            time.sleep(1)
    


def name_map(driver, text):
    while True:
        try:
            untitled_map_btn = driver.find_element_by_xpath("//div[@data-tooltip='Untitled map']")
            untitled_map_btn.click()
            print("untitled_map_btn button clicked")
            break
        except :
            time.sleep(1)

    while True:
        try:
            title_field = driver.find_element_by_xpath("//input[@type='text' and @placeholder='Untitled map']")
            title_field.clear()
            title_field.send_keys((" ".join(text.splitlines())))
            print("title field is filled.")
            print("#############" + text + "#########")
            break
        except :
            time.sleep(1)
    
    while True:
        try:
            desc_field = driver.find_element_by_xpath("//textarea[@placeholder='Add a description to help people understand your map']")
            desc_field.clear()
            desc_field.send_keys((" ".join(text.splitlines())))
            print("description field is filled.")
            break
        except :
            time.sleep(1)

    while True:
        try:
            save_btn = driver.find_element_by_xpath("//div[@id='update-map']//button[@name='save']")
            save_btn.click()
            print("save button clicked")
            break
        except :
            time.sleep(1)
        

def name_layer(driver, text):
    while True:
        try:
            untitled_layer_btn = driver.find_element_by_xpath("//div[@data-tooltip='Untitled layer']")
            untitled_layer_btn.click()
            print("untitled_layer_btn button clicked")
            break
        except :
            time.sleep(1)

    while True:
        try:
            title_field = driver.find_element_by_xpath("//input[@type='text' and @value='Untitled layer']")
            title_field.clear()
            title_field.send_keys((" ".join(text.splitlines())))
            print("title field is filled.")
            break
        except :
            time.sleep(1)

    while True:
        try:
            save_btn = driver.find_element_by_xpath("//div[@id='update-layer-name']//button[@name='save']")
            save_btn.click()
            print("save button clicked")
            break
        except :
            time.sleep(1)


def add_layer(driver):
    while True:
        try:
            add_layer_btn = driver.find_element_by_xpath("//li[@id='map-action-add-layer']")
            add_layer_btn.click()        
            print("add_layern button clicked")
            break
        except :
            time.sleep(1)
   

@fail_with_error("Cannot create pin")
def create_pin(driver, lat, lng, text): 
    global pin_count, pins_per_layer
    pin_count += 1
    if pin_count == 1: name_map(driver, text)
    if pin_count % pins_per_layer == 1: 
        if (pin_count - 1) // pins_per_layer > 0: add_layer(driver)
        name_layer(driver, text)
    print(driver.current_url) 
    url_segs = driver.current_url.split("&")
    if lat > 0 : 
        lat += 0.000005 
    else:
        lat -= 0.000005
    if lng > 0 : 
        lng += 0.000005 
    else:
        lng -= 0.000005

    url_segs[-2] = "ll=" + str(lat) + "%2C" +str(lng)
    url_segs[-1] = "z=22"
    print("&".join(url_segs))
    try:
        driver.get("&".join(url_segs))
        time.sleep(1)
        try:
            obj = driver.switch_to.alert
            obj.accept()
        except:
            pass
        time.sleep(1)

        while True:
            try:
                sel_layer = driver.find_element_by_xpath("(//div[contains(@id, 'layer-header')])[last()]/parent::div")
                sel_layer.click()
                break
            except :
                time.sleep(1)
        
        marker_click(driver)
        action = ActionChains(driver)
        elem_origin = driver.find_element_by_xpath('//html')
        print(elem_origin.rect)
        bounds = elem_origin.size
        print (bounds)
        time.sleep(2)
        
        action.move_to_element_with_offset(elem_origin, bounds['width'] / 2, bounds['height'] / 2)
        time.sleep(2)
        print(str(bounds['width'] / 2) + " , " + str(bounds['height'] / 2))
        action.click()
        print("clicked")
        action.perform()
        print("performed")
        time.sleep(2)

        while True:
            try:
                title_field = driver.find_element_by_xpath("//div[@id='map-infowindow-attr-name-value']")
                title_field.clear()
                title_field.send_keys((" ".join(text.splitlines())))
                break
            except :
                time.sleep(1)
        
        # time.sleep(2)

        while True:
            try:
                desc_field = driver.find_element_by_xpath("//div[@id='map-infowindow-attr-description-value']")
                desc_field.send_keys((" ".join(text.splitlines())))
                break
            except :
                time.sleep(1)

        while True:
            try:
                save_btn = driver.find_element_by_xpath("//div[@role='button' and text()='Save']")
                save_btn.click()
                print("save_btn clicked")
                break
            except :
                time.sleep(1)

        print("Crated a pin")
        # time.sleep(2)

    except TimeoutException:
        print("Cannot create pin")
        pass


if len(sys.argv) < 4:
    print("python 3.py [file name] [coordinate column name] [text column name]")
    exit()

file_name = sys.argv[1]
coordinate_col_name = sys.argv[2]
text_col_name = sys.argv[3]

url = 'https://accounts.google.com/'
# mail = ['janwheeler197335@gmail.com', 'aleiivrc', 'nwxdrjem']
mail_address = os.environ.get('MAIL_ADDRESS')
mail_password = os.environ.get('MAIL_PASSWORD')
mail = [mail_address, mail_password]

ua = UserAgent()
userAgent = ua.random
userAgent = userAgent.split(" ")
# userAgent[0] = "Mozilla/5.0"
userAgent = " ".join(userAgent)
print("userAgent = " + userAgent)
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('user-agent={0}'.format(userAgent))
path = '.\\webdriver\\chromedriver.exe'
chrome_options.add_argument('--log-level=0')
driver = webdriver.Chrome (executable_path = path, options = chrome_options )
# driver.maximize_window()
driver.set_window_size(1200,900)

try:
    processName["1"] = logIn
    processName["1"](mail, driver)
    time.sleep(10)
    create_map(driver)
    time.sleep(20)
    create_pins(driver, file_name, coordinate_col_name, text_col_name)
except Exception as e:
    # driver.save_screenshot(datetime.now().strftime("screenshot_%Y%m%d_%H%M%S_%f.png"))
    print(e)
finally:
    pass